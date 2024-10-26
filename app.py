# Passos manuais para automação:
# 1 - Entra na planilha e extrair o cpf do cliente
# 2 - Entra no site põe o cpf da planilha para pesquisa os status do cliente
# 3 - Verifica de esta em dia ou atrasado
# 4 - Se estiver em dia, pegar a data de pagamento e o método de pagamento
# 5 - Caso contratrio (se estiver atrasado) coloca o status como pedente
# 6 - Inserir essas novas informações (nome, cpf, vencimento, status e caso esteja em dia, data pagamento, método pagamento (cartão ou boleto)) em uma nova planilha
# 7 - Repetir ate chega um o ul
# - colar para pesquisa no path //tag[@atributo='valor']
#Vamos usar duas bibliotecas a openpyxyl para manipular planilhas e a selenium para manipular o navegador,m na casom entra no site e consuklta o cpf

import openpyxl
#permite abrir o navegador para automatizar osprocessos
from selenium import webdriver
#permite fazer a busca de elementos na pagina para que possa interagir com eles como inserir o cpf da pessoa
from selenium.webdriver.common.by import By
#Permite pausa quantos segundo quiser para deixa as coisas carregarem
from time import sleep

# 1 - Entra na planilha e extrair o cpf do cliente
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')

pagina_clientes = planilha_clientes['Sheet1']

driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')
#iter_rows uma função do pyhton para ler cada linha da planilha
for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha

    # 2 - Entra no site põe o cpf da planilha para pesquisa os status do cliente

    
    sleep(5)
    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    campo_pesquisa.clear()
    #sendkeys permite escrever uma coisa
    campo_pesquisa.send_keys(cpf)
    sleep(1)
    # 3 - Verifica de esta em dia ou atrasado
    botao_pesquisar = driver.find_element(By.XPATH, "//button[@class ='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    botao_pesquisar.click()
    sleep(4)
     
    status = driver.find_element(By.XPATH, "//span[@id ='statusLabel']")
    if status.text == "em dia":
        # 4 - Se estiver em dia, pegar a data de pagamento e o método de pagamento
        data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")
        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        #Par ELE PEGA SO OS DADOS QUE ELE QUER ELE USAR O SPLIT, essa função quebra o texto em blocos assim pode pega o bloco de acordo

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])
    
        planilha_fechamento.save('planilha fechamento.xlsx')
    else:
        # 5 - Caso contratrio (se estiver atrasado) coloca o status como pedente
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
        planilha_fechamento.save('planilha fechamento.xlsx')



   

    
    

   
